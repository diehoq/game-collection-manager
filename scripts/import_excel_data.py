#!/usr/bin/env python3
"""Import game collection and wishlist data from the provided Excel files."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook


DEFAULT_COLLECTION_FILE = "Ps2 games.xlsx"
DEFAULT_WISHLIST_FILE = "wishlist_videogiochi.xlsx"
DEFAULT_OUTPUT_FILE = "data/seed.json"


@dataclass(frozen=True)
class CollectionRecord:
    id: str
    platform: str
    title: str
    version: str
    cd_condition: str
    manual_condition: str
    price: str
    extra: str
    note: str


@dataclass(frozen=True)
class WishlistRecord:
    id: str
    platform: str
    title: str
    note: str
    in_transit: bool
    received: bool


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def normalize_bool_flag(value: object) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in {"x", "yes", "y", "true", "1", "ok"}


def normalize_header(value: object) -> str:
    text = normalize_text(value).casefold()
    text = text.replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def first_header_index(header_map: Dict[str, int], candidates: List[str]) -> int | None:
    for candidate in candidates:
        idx = header_map.get(candidate)
        if idx is not None:
            return idx
    return None


def cell_text(row: Tuple[object, ...], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return normalize_text(row[idx])


def sheet_header_map(sheet) -> Dict[str, int]:
    if sheet.max_row < 1:
        return {}
    values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping: Dict[str, int] = {}
    for idx, value in enumerate(values):
        key = normalize_header(value)
        if key:
            mapping[key] = idx
    return mapping


def build_collection_key(platform: str, title: str) -> Tuple[str, str]:
    return (platform.strip().casefold(), title.strip().casefold())


def parse_collection_sheet(path: Path) -> List[CollectionRecord]:
    workbook = load_workbook(path, data_only=True)
    rows: List[CollectionRecord] = []
    seen_keys = set()
    next_id = 1

    for sheet in workbook.worksheets:
        header_map = sheet_header_map(sheet)
        title_idx = first_header_index(header_map, ["game", "title", "titolo"])
        if title_idx is None:
            continue

        platform_idx = first_header_index(header_map, ["platform", "piattaforma"])
        version_idx = first_header_index(header_map, ["version"])
        cd_idx = first_header_index(header_map, ["cd", "disc", "cd condition", "disc condition"])
        manual_idx = first_header_index(header_map, ["manual", "manual condition"])
        price_idx = first_header_index(header_map, ["price", "prezzo"])
        extra_idx = first_header_index(header_map, ["extra"])
        note_idx = first_header_index(header_map, ["note", "notes", "note:"])

        default_platform = normalize_text(sheet.title)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            title = cell_text(row, title_idx)
            if not title:
                continue

            platform = cell_text(row, platform_idx) or default_platform
            key = build_collection_key(platform, title)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            record = CollectionRecord(
                id=f"c{next_id}",
                platform=platform,
                title=title,
                version=cell_text(row, version_idx),
                cd_condition=cell_text(row, cd_idx),
                manual_condition=cell_text(row, manual_idx),
                price=cell_text(row, price_idx),
                extra=cell_text(row, extra_idx),
                note=cell_text(row, note_idx),
            )
            rows.append(record)
            next_id += 1

    return rows


def parse_wishlist_sheet(path: Path) -> List[WishlistRecord]:
    workbook = load_workbook(path, data_only=True)
    rows: List[WishlistRecord] = []
    next_id = 1

    for sheet in workbook.worksheets:
        header_map = sheet_header_map(sheet)
        title_idx = first_header_index(header_map, ["titolo", "title", "game"])
        if title_idx is None:
            continue

        platform_idx = first_header_index(header_map, ["platform", "piattaforma"])
        note_idx = first_header_index(header_map, ["note", "notes", "note:"])
        received_idx = first_header_index(header_map, ["acquistato", "received", "ricevuto"])
        in_transit_idx = first_header_index(
            header_map, ["in transito", "in transit", "in-transit", "transit"]
        )

        default_platform = normalize_text(sheet.title)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            title = cell_text(row, title_idx)
            if not title:
                continue
            note = cell_text(row, note_idx)
            received = normalize_bool_flag(cell_text(row, received_idx))
            in_transit = normalize_bool_flag(cell_text(row, in_transit_idx))
            platform = cell_text(row, platform_idx) or default_platform
            rows.append(
                WishlistRecord(
                    id=f"w{next_id}",
                    platform=platform,
                    title=title,
                    note=note,
                    in_transit=in_transit,
                    received=received,
                )
            )
            next_id += 1

    return rows


def transfer_received_items(
    collection_records: List[CollectionRecord], wishlist_records: List[WishlistRecord]
) -> Tuple[List[CollectionRecord], List[WishlistRecord]]:
    collection_by_key: Dict[Tuple[str, str], CollectionRecord] = {
        build_collection_key(item.platform, item.title): item for item in collection_records
    }
    migrated_collection = list(collection_records)
    migrated_wishlist: List[WishlistRecord] = []

    next_collection_id = len(migrated_collection) + 1
    for wish in wishlist_records:
        if wish.received:
            key = build_collection_key(wish.platform, wish.title)
            if key not in collection_by_key:
                note = wish.note
                migrated = CollectionRecord(
                    id=f"c{next_collection_id}",
                    platform=wish.platform,
                    title=wish.title,
                    version="",
                    cd_condition="",
                    manual_condition="",
                    price="",
                    extra="",
                    note=note,
                )
                migrated_collection.append(migrated)
                collection_by_key[key] = migrated
                next_collection_id += 1
            continue
        migrated_wishlist.append(wish)

    return migrated_collection, migrated_wishlist


def sorted_records(
    records: Iterable[CollectionRecord | WishlistRecord],
) -> List[CollectionRecord | WishlistRecord]:
    return sorted(records, key=lambda r: (r.platform.casefold(), r.title.casefold()))


def to_json_payload(
    collection_records: List[CollectionRecord], wishlist_records: List[WishlistRecord]
) -> dict:
    collection_json = [
        {
            "id": row.id,
            "platform": row.platform,
            "title": row.title,
            "version": row.version,
            "cdCondition": row.cd_condition,
            "manualCondition": row.manual_condition,
            "price": row.price,
            "extra": row.extra,
            "note": row.note,
        }
        for row in sorted_records(collection_records)
    ]
    wishlist_json = [
        {
            "id": row.id,
            "platform": row.platform,
            "title": row.title,
            "note": row.note,
            "inTransit": row.in_transit,
            "received": row.received,
        }
        for row in sorted_records(wishlist_records)
    ]

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "collection": collection_json,
        "wishlist": wishlist_json,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import collection and wishlist spreadsheets into JSON seed data."
    )
    parser.add_argument(
        "--collection",
        default=DEFAULT_COLLECTION_FILE,
        help="Path to collection .xlsx file.",
    )
    parser.add_argument(
        "--wishlist",
        default=DEFAULT_WISHLIST_FILE,
        help="Path to wishlist .xlsx file.",
    )
    parser.add_argument(
        "--out",
        default=DEFAULT_OUTPUT_FILE,
        help="Output JSON file path.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    collection_path = Path(args.collection)
    wishlist_path = Path(args.wishlist)
    out_path = Path(args.out)

    if not collection_path.exists():
        raise FileNotFoundError(f"Collection file not found: {collection_path}")
    if not wishlist_path.exists():
        raise FileNotFoundError(f"Wishlist file not found: {wishlist_path}")

    collection_rows = parse_collection_sheet(collection_path)
    wishlist_rows = parse_wishlist_sheet(wishlist_path)
    collection_rows, wishlist_rows = transfer_received_items(collection_rows, wishlist_rows)
    payload = to_json_payload(collection_rows, wishlist_rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {out_path} ({len(collection_rows)} collection items, {len(wishlist_rows)} wishlist items)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
