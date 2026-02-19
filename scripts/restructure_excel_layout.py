#!/usr/bin/env python3
"""Restructure workbook layouts into a single-table, filter-friendly format."""

from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from import_excel_data import (
    DEFAULT_COLLECTION_FILE,
    DEFAULT_WISHLIST_FILE,
    CollectionRecord,
    WishlistRecord,
    parse_collection_sheet,
    parse_wishlist_sheet,
    transfer_received_items,
)


PLATFORMS = ["PS1", "PS2", "PS4", "DS WII"]
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TRANSIT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
RECEIVED_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Restructure collection and wishlist Excel files into functional layouts."
    )
    parser.add_argument("--collection", default=DEFAULT_COLLECTION_FILE, help="Collection workbook path.")
    parser.add_argument("--wishlist", default=DEFAULT_WISHLIST_FILE, help="Wishlist workbook path.")
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Overwrite files without creating timestamped backups first.",
    )
    return parser.parse_args()


def autosize_columns(ws) -> None:
    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))


def style_sheet_header(ws, column_count: int) -> None:
    for cell in ws[1][:column_count]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_table(ws, name: str) -> None:
    if ws.max_row < 2:
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table_style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = table_style
    ws.add_table(table)


def ordered_collection(records: Iterable[CollectionRecord]) -> List[CollectionRecord]:
    return sorted(records, key=lambda r: (r.platform.casefold(), r.title.casefold()))


def ordered_wishlist(records: Iterable[WishlistRecord]) -> List[WishlistRecord]:
    return sorted(records, key=lambda r: (r.platform.casefold(), r.title.casefold()))


def build_collection_workbook(records: List[CollectionRecord]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"
    headers = ["ID", "Platform", "Title", "Version", "CD", "Manual", "Price", "Extra", "Note"]
    ws.append(headers)

    for idx, row in enumerate(ordered_collection(records), start=1):
        ws.append(
            [
                f"C{idx:04d}",
                row.platform,
                row.title,
                row.version,
                row.cd_condition,
                row.manual_condition,
                row.price,
                row.extra,
                row.note,
            ]
        )

    style_sheet_header(ws, len(headers))
    add_table(ws, "CollectionTable")

    platform_validation = DataValidation(type="list", formula1='"PS1,PS2,PS4,DS WII"', allow_blank=True)
    ws.add_data_validation(platform_validation)
    if ws.max_row >= 2:
        platform_validation.add(f"B2:B{ws.max_row}")

    autosize_columns(ws)
    add_collection_summary_sheet(wb)
    return wb


def add_collection_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Total games", "=COUNTA(Collection!C:C)-1"])
    for platform in PLATFORMS:
        ws.append([f"{platform} games", f'=COUNTIF(Collection!B:B,"{platform}")'])
    style_sheet_header(ws, 2)
    autosize_columns(ws)


def build_wishlist_workbook(records: List[WishlistRecord]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wishlist"
    headers = [
        "ID",
        "Platform",
        "Title",
        "Note",
        "Priority",
        "In Transit",
        "Received",
        "Ordered Date",
        "Received Date",
    ]
    ws.append(headers)

    for idx, row in enumerate(ordered_wishlist(records), start=1):
        ws.append(
            [
                f"W{idx:04d}",
                row.platform,
                row.title,
                row.note,
                "Medium",
                "Yes" if row.in_transit else "No",
                "No",
                "",
                "",
            ]
        )

    style_sheet_header(ws, len(headers))
    add_table(ws, "WishlistTable")
    add_wishlist_validations(ws)
    add_wishlist_conditional_formatting(ws)
    autosize_columns(ws)
    add_wishlist_summary_sheet(wb)
    return wb


def add_wishlist_validations(ws) -> None:
    if ws.max_row < 2:
        return
    platform_validation = DataValidation(type="list", formula1='"PS1,PS2,PS4,DS WII"', allow_blank=True)
    yes_no_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    priority_validation = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
    ws.add_data_validation(platform_validation)
    ws.add_data_validation(yes_no_validation)
    ws.add_data_validation(priority_validation)

    platform_validation.add(f"B2:B{ws.max_row}")
    priority_validation.add(f"E2:E{ws.max_row}")
    yes_no_validation.add(f"F2:G{ws.max_row}")


def add_wishlist_conditional_formatting(ws) -> None:
    if ws.max_row < 2:
        return
    row_range = f"A2:I{ws.max_row}"
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=['=$F2="Yes"'], stopIfTrue=False, fill=TRANSIT_FILL),
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=['=$G2="Yes"'], stopIfTrue=False, fill=RECEIVED_FILL),
    )


def add_wishlist_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Pending wishlist items", '=COUNTIFS(Wishlist!G:G,"No")'])
    ws.append(["In transit items", '=COUNTIFS(Wishlist!F:F,"Yes",Wishlist!G:G,"No")'])
    ws.append(["Received (to move)", '=COUNTIFS(Wishlist!G:G,"Yes")'])
    for platform in PLATFORMS:
        ws.append([f"{platform} pending", f'=COUNTIFS(Wishlist!B:B,"{platform}",Wishlist!G:G,"No")'])
    style_sheet_header(ws, 2)
    autosize_columns(ws)


def backup_path(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")


def main() -> int:
    args = parse_args()
    collection_path = Path(args.collection)
    wishlist_path = Path(args.wishlist)

    if not collection_path.exists():
        raise FileNotFoundError(f"Collection file not found: {collection_path}")
    if not wishlist_path.exists():
        raise FileNotFoundError(f"Wishlist file not found: {wishlist_path}")

    collection_records = parse_collection_sheet(collection_path)
    wishlist_records = parse_wishlist_sheet(wishlist_path)
    collection_records, wishlist_records = transfer_received_items(collection_records, wishlist_records)

    backups: List[Path] = []
    if not args.no_backup:
        collection_backup = backup_path(collection_path)
        wishlist_backup = backup_path(wishlist_path)
        shutil.copy2(collection_path, collection_backup)
        shutil.copy2(wishlist_path, wishlist_backup)
        backups.extend([collection_backup, wishlist_backup])

    collection_wb = build_collection_workbook(collection_records)
    wishlist_wb = build_wishlist_workbook(wishlist_records)
    collection_wb.save(collection_path)
    wishlist_wb.save(wishlist_path)

    print(f"Restructured {collection_path} ({len(collection_records)} items)")
    print(f"Restructured {wishlist_path} ({len(wishlist_records)} items)")
    if backups:
        for backup in backups:
            print(f"Backup created: {backup}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
